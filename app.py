import os
import json
import sqlite3
import datetime
import io
import csv
try:
    from zoneinfo import ZoneInfo          # Python 3.9+
except ImportError:
    from backports.zoneinfo import ZoneInfo  # Python 3.8 fallback

JST = ZoneInfo('Asia/Tokyo')

def now_jst() -> datetime.datetime:
    """現在の日本時間（JST）を返す。"""
    return datetime.datetime.now(tz=JST)

def utc_str_to_jst_str(utc_str: str) -> str:
    """
    SQLite の CURRENT_TIMESTAMP 形式（'YYYY-MM-DD HH:MM:SS'、UTC）を
    JST の同形式文字列に変換する。
    すでに JST で保存済みの値はそのまま返す。
    """
    if not utc_str:
        return utc_str
    try:
        # SQLite は TZ なし文字列で UTC を返す
        dt_utc = datetime.datetime.fromisoformat(utc_str).replace(
            tzinfo=datetime.timezone.utc
        )
        return dt_utc.astimezone(JST).strftime('%Y-%m-%d %H:%M:%S')
    except ValueError:
        return utc_str
import openpyxl
from utils.stress_text import generate_advice
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, make_response

# ==========================================
# ストレージバックエンド切り替え
# USE_FIRESTORE=1 → Cloud Firestore（本番）
# USE_FIRESTORE=0 / 未設定 → SQLite（ローカル開発）
# ==========================================
USE_FIRESTORE = os.environ.get('USE_FIRESTORE', '0') == '1'

if USE_FIRESTORE:
    from google.cloud import firestore as _fs
    _fdb = _fs.Client()

    @_fs.transactional
    def _txn_save_response(transaction, counter_ref, doc_ref, doc_data):
        """
        Firestoreトランザクション内で:
          1. metadata/counters の last_employee_id を +1 して取得（連番採番）
          2. responses コレクションに employee_id 付きでドキュメントを保存
        """
        snap = counter_ref.get(transaction=transaction)
        new_id = (snap.get('last_employee_id') + 1) if snap.exists else 1
        if snap.exists:
            transaction.update(counter_ref, {'last_employee_id': new_id})
        else:
            transaction.set(counter_ref, {'last_employee_id': new_id})
        transaction.set(doc_ref, {**doc_data, 'employee_id': str(new_id)})
        return str(new_id)

# ── ストレージ抽象化関数 ─────────────────────────────────────────────────────

def _db_save_response(name, workplace_name, email, is_high_stress, answers_json) -> str:
    """回答を保存して doc_id（文字列）を返す。"""
    jst_now = now_jst().strftime('%Y-%m-%d %H:%M:%S')
    if USE_FIRESTORE:
        doc_data = {
            'name': name, 'workplace_name': workplace_name or '',
            'email': email, 'is_high_stress': bool(is_high_stress),
            'answers_json': answers_json, 'created_at': jst_now,
        }
        counter_ref = _fdb.collection('metadata').document('counters')
        new_doc_ref = _fdb.collection('responses').document()
        # トランザクションで連番採番 + 保存を原子的に実行
        _txn_save_response(_fdb.transaction(), counter_ref, new_doc_ref, doc_data)
        return new_doc_ref.id
    else:
        conn = get_db_connection()
        cursor = conn.execute(
            'INSERT INTO responses (employee_id,name,workplace_name,email,is_high_stress,answers_json,created_at) VALUES (?,?,?,?,?,?,?)',
            ('', name, workplace_name, email, is_high_stress, answers_json, jst_now)
        )
        row_id = cursor.lastrowid
        conn.execute('UPDATE responses SET employee_id=? WHERE id=?', (str(row_id), row_id))
        conn.commit()
        conn.close()
        return str(row_id)

def _db_get_response(doc_id: str) -> dict:
    """doc_id に対応する回答を dict で返す。見つからなければ None。"""
    if USE_FIRESTORE:
        doc = _fdb.collection('responses').document(doc_id).get()
        if not doc.exists:
            return None
        d = doc.to_dict()
        d['id'] = doc.id
        return d
    else:
        conn = get_db_connection()
        try:
            row = conn.execute('SELECT * FROM responses WHERE id=?', (int(doc_id),)).fetchone()
        except (ValueError, TypeError):
            row = None
        conn.close()
        return dict(row) if row else None

def _db_list_responses() -> list:
    """全回答を created_at 降順で返す（list of dict）。"""
    if USE_FIRESTORE:
        docs = _fdb.collection('responses').order_by(
            'created_at', direction=_fs.Query.DESCENDING
        ).stream()
        result = []
        for doc in docs:
            d = doc.to_dict()
            d['id'] = doc.id
            result.append(d)
        return result
    else:
        conn = get_db_connection()
        rows = conn.execute('SELECT * FROM responses ORDER BY created_at DESC').fetchall()
        conn.close()
        result = []
        for r in rows:
            d = dict(r)
            d['id'] = str(d['id'])  # Firestore と型を統一（文字列）
            result.append(d)
        return result

def _db_get_setting(key, default='') -> str:
    """設定値を取得する。"""
    if USE_FIRESTORE:
        doc = _fdb.collection('settings').document(key).get()
        return doc.to_dict().get('value', default) if doc.exists else default
    else:
        conn = get_db_connection()
        row = conn.execute('SELECT value FROM settings WHERE key=?', (key,)).fetchone()
        conn.close()
        return row['value'] if row else default

def _db_list_settings() -> dict:
    """全設定を dict で返す。"""
    if USE_FIRESTORE:
        docs = _fdb.collection('settings').stream()
        return {doc.id: doc.to_dict().get('value', '') for doc in docs}
    else:
        conn = get_db_connection()
        rows = conn.execute('SELECT key, value FROM settings').fetchall()
        conn.close()
        return {row['key']: row['value'] for row in rows}

def _db_save_settings(settings_dict: dict):
    """複数の設定をまとめて保存する。"""
    if USE_FIRESTORE:
        batch = _fdb.batch()
        for key, value in settings_dict.items():
            ref = _fdb.collection('settings').document(str(key))
            batch.set(ref, {'value': str(value)})
        batch.commit()
    else:
        conn = get_db_connection()
        for key, value in settings_dict.items():
            conn.execute('INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)',
                         (key, str(value)))
        conn.commit()
        conn.close()

def _db_list_workplaces() -> list:
    """全職場を code 昇順で返す（list of dict）。"""
    if USE_FIRESTORE:
        docs = _fdb.collection('workplaces').order_by('code').stream()
        result = []
        for doc in docs:
            d = doc.to_dict()
            d['id'] = doc.id
            result.append(d)
        return result
    else:
        conn = get_db_connection()
        rows = conn.execute('SELECT * FROM workplaces ORDER BY code').fetchall()
        conn.close()
        return [dict(r) for r in rows]

def _db_add_workplace(code, name, level) -> str:
    """職場を追加して doc_id を返す。"""
    if USE_FIRESTORE:
        _, doc_ref = _fdb.collection('workplaces').add(
            {'code': code, 'name': name, 'level': int(level)}
        )
        return doc_ref.id
    else:
        conn = get_db_connection()
        cursor = conn.execute(
            'INSERT INTO workplaces (code,name,level) VALUES (?,?,?)', (code, name, level)
        )
        wp_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return str(wp_id)

def _db_delete_workplace(doc_id: str):
    """職場を削除する。"""
    if USE_FIRESTORE:
        _fdb.collection('workplaces').document(doc_id).delete()
    else:
        conn = get_db_connection()
        try:
            conn.execute('DELETE FROM workplaces WHERE id=?', (int(doc_id),))
        except (ValueError, TypeError):
            pass
        conn.commit()
        conn.close()

app = Flask(__name__)
# セッション（結果画面へのデータ受け渡し）を使うための暗号化キー
app.secret_key = 'stress_check_super_secret_key'

# ==========================================
# 素点換算表（MHLW公式・性別対応）
# ==========================================
_CT = {
    "male": {
        "A1_quantity":      [(3,5,1),(6,7,2),(8,9,3),(10,11,4),(12,12,5)],
        "A2_quality":       [(3,5,1),(6,7,2),(8,9,3),(10,11,4),(12,12,5)],
        "A3_physical":      [(1,1,2),(2,2,3),(3,3,4),(4,4,5)],
        "A4_interpersonal": [(3,3,1),(4,5,2),(6,7,3),(8,9,4),(10,12,5)],
        "A5_environment":   [(1,1,2),(2,2,3),(3,3,4),(4,4,5)],
        "A6_control":       [(3,4,1),(5,6,2),(7,8,3),(9,10,4),(11,12,5)],
        "A7_skill":         [(1,1,1),(2,2,2),(3,3,3),(4,4,4)],
        "A8_suitability":   [(1,1,1),(2,2,2),(3,3,3),(4,4,5)],
        "A9_reward":        [(1,1,1),(2,2,2),(3,3,3),(4,4,5)],
        "B1_vigor":         [(3,3,1),(4,5,2),(6,7,3),(8,9,4),(10,12,5)],
        "B2_irritation":    [(3,3,1),(4,5,2),(6,7,3),(8,9,4),(10,12,5)],
        "B3_fatigue":       [(3,3,1),(4,4,2),(5,7,3),(8,10,4),(11,12,5)],
        "B4_anxiety":       [(3,3,1),(4,4,2),(5,7,3),(8,9,4),(10,12,5)],
        "B5_depression":    [(6,6,1),(7,8,2),(9,12,3),(13,16,4),(17,24,5)],
        "B6_physical":      [(11,11,1),(12,15,2),(16,21,3),(22,26,4),(27,44,5)],
        "C1_boss":          [(3,4,1),(5,6,2),(7,8,3),(9,10,4),(11,12,5)],
        "C2_coworker":      [(3,5,1),(6,7,2),(8,9,3),(10,11,4),(12,12,5)],
        "C3_family":        [(3,6,1),(7,8,2),(9,9,3),(10,11,4),(12,12,5)],
        "D1_satisfaction":  [(2,3,1),(4,4,2),(5,6,3),(7,7,4),(8,8,5)],
    },
    "female": {
        "A1_quantity":      [(3,4,1),(5,6,2),(7,9,3),(10,11,4),(12,12,5)],
        "A2_quality":       [(3,4,1),(5,6,2),(7,8,3),(9,10,4),(11,12,5)],
        "A3_physical":      [(1,1,2),(2,2,3),(3,3,4),(4,4,5)],
        "A4_interpersonal": [(3,3,1),(4,5,2),(6,7,3),(8,9,4),(10,12,5)],
        "A5_environment":   [(1,1,1),(2,2,3),(3,3,4),(4,4,5)],
        "A6_control":       [(3,3,1),(4,5,2),(6,8,3),(9,10,4),(11,12,5)],
        "A7_skill":         [(1,1,1),(2,2,2),(3,3,3),(4,4,4)],
        "A8_suitability":   [(1,1,1),(2,2,2),(3,3,3),(4,4,5)],
        "A9_reward":        [(1,1,1),(2,2,2),(3,3,3),(4,4,5)],
        "B1_vigor":         [(3,3,1),(4,5,2),(6,7,3),(8,9,4),(10,12,5)],
        "B2_irritation":    [(3,3,1),(4,5,2),(6,8,3),(9,10,4),(11,12,5)],
        "B3_fatigue":       [(3,3,1),(4,5,2),(6,8,3),(9,11,4),(12,12,5)],
        "B4_anxiety":       [(3,3,1),(4,4,2),(5,7,3),(8,10,4),(11,12,5)],
        "B5_depression":    [(6,6,1),(7,8,2),(9,12,3),(13,17,4),(18,24,5)],
        "B6_physical":      [(11,13,1),(14,17,2),(18,23,3),(24,29,4),(30,44,5)],
        "C1_boss":          [(3,3,1),(4,5,2),(6,7,3),(8,10,4),(11,12,5)],
        "C2_coworker":      [(3,5,1),(6,7,2),(8,9,3),(10,11,4),(12,12,5)],
        "C3_family":        [(3,6,1),(7,8,2),(9,9,3),(10,11,4),(12,12,5)],
        "D1_satisfaction":  [(2,3,1),(4,4,2),(5,6,3),(7,7,4),(8,8,5)],
    }
}

def _get_ep(gender_key, subscale, raw):
    for mn, mx, pt in _CT[gender_key].get(subscale, []):
        if mn <= raw <= mx:
            return pt
    return 3

# ==========================================
# データベース(SQLite)の初期設定
# ==========================================
DB_NAME = 'database.db'

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    # 職場（組織図）テーブル
    c.execute('''
        CREATE TABLE IF NOT EXISTS workplaces (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL,
            name TEXT NOT NULL,
            level INTEGER NOT NULL
        )
    ''')
    # 設定テーブル（管理画面の実施者管理と連動）
    c.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL DEFAULT ''
        )
    ''')
    # 回答結果テーブル（回答内容は扱いやすいようにJSONで丸ごと保存）
    c.execute('''
        CREATE TABLE IF NOT EXISTS responses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id TEXT NOT NULL,
            name TEXT NOT NULL,
            workplace_name TEXT,
            email TEXT,
            is_high_stress BOOLEAN,
            answers_json TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

# アプリ起動時にDBを作成
init_db()

def migrate_db():
    """既存DBに不足カラムを追加するマイグレーション処理"""
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    # responses テーブルに email カラムが存在しない場合は追加
    existing_cols = [row[1] for row in c.execute('PRAGMA table_info(responses)').fetchall()]
    if 'email' not in existing_cols:
        c.execute('ALTER TABLE responses ADD COLUMN email TEXT')
    conn.commit()
    conn.close()

migrate_db()

def get_db_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

def get_setting(key, default=''):
    return _db_get_setting(key, default)

# ==========================================
# ストレスチェック 判定 ＆ 詳細スコア計算ロジック
# ==========================================
def analyze_stress(answers):
    gender = answers.get('gender', '男性')
    g = 'female' if gender == '女性' else 'male'
    q = {f'q{i}': int(answers.get(f'q{i}', 3)) for i in range(1, 58)}

    # 1. 素点計算（MHLW 57問マッピング）
    raw = {
        'A1_quantity':      15 - (q['q1'] + q['q2'] + q['q3']),
        'A2_quality':       15 - (q['q4'] + q['q5'] + q['q6']),
        'A3_physical':       5 -  q['q7'],
        'A4_interpersonal': 10 - (q['q12'] + q['q13']) + q['q14'],
        'A5_environment':    5 -  q['q15'],
        'A6_control':       15 - (q['q8']  + q['q9']  + q['q10']),
        'A7_skill':               q['q11'],
        'A8_suitability':    5 -  q['q16'],
        'A9_reward':         5 -  q['q17'],
        'B1_vigor':          q['q18'] + q['q19'] + q['q20'],
        'B2_irritation':     q['q21'] + q['q22'] + q['q23'],
        'B3_fatigue':        q['q24'] + q['q25'] + q['q26'],
        'B4_anxiety':        q['q27'] + q['q28'] + q['q29'],
        'B5_depression':     q['q30'] + q['q31'] + q['q32'] + q['q33'] + q['q34'] + q['q35'],
        'B6_physical':       q['q36'] + q['q37'] + q['q38'] + q['q39'] + q['q40'] + q['q41'] +
                             q['q42'] + q['q43'] + q['q44'] + q['q45'] + q['q46'],
        'C1_boss':     15 - (q['q47'] + q['q50'] + q['q53']),
        'C2_coworker': 15 - (q['q48'] + q['q51'] + q['q54']),
        'C3_family':   15 - (q['q49'] + q['q52'] + q['q55']),
        'D1_satisfaction': 10 - (q['q56'] + q['q57']),
    }

    # 2. 評価点（ep）に変換
    ep = {key: _get_ep(g, key, val) for key, val in raw.items()}

    # 3. ヘルススコア合計（フロントエンド表示と完全一致させる）
    # STAR軸（★）: ep をそのまま使用（高ep=良好）
    # 非STAR軸（非★）: 6-ep に変換（低ep=良好 → 外周=良好に統一）
    # この変換により「合計点が高いほど良い状態」に統一される
    _STAR = frozenset({'A6_control','A7_skill','A8_suitability','A9_reward',
                       'B1_vigor','C1_boss','C2_coworker','C3_family','D1_satisfaction'})
    def _h(k): return ep[k] if k in _STAR else (6 - ep[k])

    sumA = (_h('A1_quantity') + _h('A2_quality') + _h('A3_physical') +
            _h('A4_interpersonal') + _h('A5_environment') +
            _h('A6_control') + _h('A7_skill') + _h('A8_suitability') + _h('A9_reward'))
    sumB = (_h('B1_vigor') + _h('B2_irritation') + _h('B3_fatigue') +
            _h('B4_anxiety') + _h('B5_depression') + _h('B6_physical'))
    sumC = _h('C1_boss') + _h('C2_coworker') + _h('C3_family')  # D1は除外

    # 4. 高ストレス判定（厚労省方式 素点換算表を用いた評価基準 その2）
    # 条件①: 領域B合計（健康スコア合計, 6-30点）≤ 12
    # 条件②: 領域B合計 ≤ 17 かつ (領域A + 領域C)合計 ≤ 26
    is_high_stress = (sumB <= 12) or (sumB <= 17 and (sumA + sumC) <= 26)

    # 2. 5段階スケール変換（1:悪い ～ 5:良い）
    def to_5_scale(q_list, reverse_goodness=False):
        total = sum([(5 - int(answers.get(f'q{i}', 3))) if reverse_goodness else int(answers.get(f'q{i}', 3)) for i in q_list])
        avg = total / len(q_list)
        return round(1 + (avg - 1) * (4 / 3), 1)

    # レーダーチャート用（7軸）
    radar_scores = [
        to_5_scale([1,2,3,4,5,6,7], False),     # 負担
        to_5_scale([8,9,10], True),             # コントロール
        to_5_scale([18,19,20], False),          # 活気
        to_5_scale([21,22,23,24,25,26], True),  # イライラ
        to_5_scale([27,28,29,30,31,32,33,34,35], True), # 不安・抑うつ
        to_5_scale([47,50,53], True),           # 上司支援
        to_5_scale([48,51,54], True)            # 同僚支援
    ]

    # 3. 詳細棒グラフ用スコア ＆ 自動コメント生成
    domain_a_avg = to_5_scale(list(range(1, 18)), False)
    domain_b_avg = to_5_scale(list(range(18, 47)), True)
    domain_c_avg = to_5_scale(list(range(47, 56)), True)

    advice_text = "日々の業務お疲れ様です。"
    if domain_b_avg <= 2.5:
        advice_text += "心身のストレス反応が強く出ているようです。休息を最優先にしてください。"
    elif domain_a_avg <= 2.5:
        advice_text += "仕事の量や質による負担が大きいようです。業務の調整や周囲への相談を検討してみましょう。"
    elif domain_c_avg <= 2.5:
        advice_text += "周囲のサポートが不足していると感じているようです。一人で抱え込まず、発信することが大切です。"
    else:
        advice_text += "現在のところ、全体的なバランスは比較的良好に保たれています。引き続きセルフケアを継続してください。"

    # 個別アドバイス生成（ep合計はA/B/Cそれぞれ単純合計）
    advice_detail = generate_advice(ep, sumA, sumB, sumC, is_high_stress,
                                    name=answers.get('name', ''))

    return {
        'is_high_stress': is_high_stress,
        'radar_scores': radar_scores,
        'bars': {
            'domain_a': domain_a_avg,
            'domain_b': domain_b_avg,
            'domain_c': domain_c_avg
        },
        'advice_text': advice_text,
        'advice_detail': advice_detail,
    }

# ==========================================
# ルーティング（画面表示）
# ==========================================

@app.route('/')
def index():
    # 従業員ページ：DBから職場一覧と設定を取得して渡す
    workplaces = _db_list_workplaces()
    company_name      = get_setting('company_name', 'デモ企業')
    practitioner_name = get_setting('practitioner_name', '（未設定）')
    practitioner_role = get_setting('practitioner_role', '')
    return render_template('index.html',
                           company_name=company_name,
                           practitioner_name=practitioner_name,
                           practitioner_role=practitioner_role,
                           workplaces=workplaces)

@app.route('/admin')
def admin():
    # 管理者ダッシュボード
    return render_template('admin.html')

@app.route('/result')
def show_result():
    """
    従業員の個人結果レポート画面。
    セッションは使用しない（クッキーサイズ上限問題を回避）。
    URL パラメータ result_id で DB レコードを特定し、
    analyze_stress() を再実行してテンプレートに渡す。
    """
    # result_id は SQLite では数字文字列、Firestore では英数字文字列
    result_id = request.args.get('result_id')
    if not result_id:
        return redirect(url_for('index'))

    # DB からレコードを1件取得
    row = _db_get_response(result_id)
    if not row:
        return redirect(url_for('index'))

    # answers_json からフォーム送信時のデータを復元
    try:
        submitted = json.loads(row['answers_json'])
    except Exception:
        submitted = {}

    # スコア・アドバイスを再計算（DB 保存済みデータから確実に復元）
    analysis = analyze_stress(submitted)

    # フロント React 用に q1〜q57 の生回答だけ抽出
    raw_answers = {f'q{i}': int(submitted.get(f'q{i}', 3)) for i in range(1, 58)}

    # 保存日時（JST で保存済み、念のため変換も通す）
    created_at = row.get('created_at') or ''
    exam_date = utc_str_to_jst_str(created_at)[:10] if created_at else now_jst().strftime('%Y-%m-%d')

    return render_template('result.html',
        employee_id   = row.get('employee_id', result_id),
        name          = row.get('name', ''),
        workplace_name= row.get('workplace_name', ''),
        exam_date     = exam_date,
        ref_number    = f"SC-{result_id}",
        gender        = submitted.get('gender', ''),
        is_high_stress= analysis['is_high_stress'],
        radar_scores  = analysis['radar_scores'],
        bars          = analysis['bars'],
        advice_text   = analysis['advice_text'],
        advice_detail = analysis['advice_detail'],
        answers_json  = json.dumps(raw_answers, ensure_ascii=False),
    )

# ==========================================
# APIエンドポイント（データ処理）
# ==========================================

@app.route('/api/submit', methods=['POST'])
def submit_data():
    """ 従業員からの回答データを受け取る """
    data = request.json

    # 1. 高ストレス判定とスコア計算を実行
    analysis = analyze_stress(data)

    # 2. データベースに保存（SQLite or Firestore）
    row_id = _db_save_response(
        name           = data.get('name', ''),
        workplace_name = data.get('workplace_name', ''),
        email          = data.get('email', '') or None,
        is_high_stress = analysis['is_high_stress'],
        answers_json   = json.dumps(data, ensure_ascii=False),
    )

    # 3. フロントエンドにリダイレクト先を返す
    # セッションは使用しない（クッキーサイズ上限問題を回避）。
    # result_id を URL に含めるだけで、show_result() が DB から再構築する。
    return jsonify({'success': True, 'redirect_url': url_for('show_result', result_id=row_id)})

# ----------------- 職場管理API -----------------
@app.route('/api/workplaces', methods=['GET', 'POST'])
def manage_workplaces():
    if request.method == 'POST':
        data = request.json
        _db_add_workplace(data['code'], data['name'], data['level'])
        return jsonify({'success': True})
    return jsonify(_db_list_workplaces())

@app.route('/api/workplaces/<doc_id>', methods=['DELETE'])
def delete_workplace(doc_id):
    _db_delete_workplace(doc_id)
    return jsonify({'success': True})

# ----------------- 結果一覧取得API（管理者画面用） -----------------
@app.route('/api/responses')
def get_responses():
    rows = _db_list_responses()
    results = []
    for r in rows:
        try:
            ans = json.loads(r.get('answers_json') or '{}')
        except Exception:
            ans = {}
        # created_at は JST で保存済み。念のため変換も通す
        submitted_at = utc_str_to_jst_str(r.get('created_at') or '')
        results.append({
            'id':             r.get('id'),
            'employee_id':    r.get('employee_id', ''),
            'name':           r.get('name', ''),
            'furigana':       ans.get('furigana', ''),
            'birth_date':     ans.get('birth_date', ''),
            'gender':         ans.get('gender', ''),
            'workplace_code': ans.get('workplace_code', ''),
            'workplace_name': r.get('workplace_name', ''),
            'is_high_stress': r.get('is_high_stress', False),
            'submitted_at':   submitted_at,
        })
    return jsonify(results)

# ----------------- 設定管理API（管理画面の実施者管理と連動） -----------------
@app.route('/api/settings', methods=['GET', 'POST'])
def manage_settings():
    if request.method == 'POST':
        _db_save_settings(request.json or {})
        return jsonify({'success': True})
    return jsonify(_db_list_settings())

# ----------------- Excel出力API（職場データ・取込用フォーマット） -----------------
@app.route('/api/csv/workplaces')
def export_workplaces_csv():
    workplaces = _db_list_workplaces()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'newSheet'

    # ヘッダー行の定義（添付Excelと同じ13列）
    headers = [
        '職場コード', '前回職場コード',
        '第1職場名', '第2職場名', '第3職場名', '第4職場名', '第5職場名',
        '第6職場名', '第7職場名', '第8職場名', '第9職場名', '第10職場名',
        '初回パスワード'
    ]

    # ヘッダースタイル
    header_fill   = PatternFill('solid', start_color='1F497D', end_color='1F497D')
    header_font   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # データ行：level に応じて第N職場名カラムに名前を入れる
    data_font  = Font(name='Arial', size=10)
    data_align = Alignment(vertical='center')

    for row_idx, wp in enumerate(workplaces, start=2):
        level = int(wp.get('level') or 1)
        # 13列分の空行を作成し、職場コードと該当する第N職場名だけ埋める
        row_data = [''] * 13
        row_data[0] = str(wp.get('code', ''))   # A: 職場コード
        row_data[1] = ''                         # B: 前回職場コード（空）
        name_col = 1 + level                     # C=第1(level1), D=第2(level2)...
        if 2 <= name_col <= 11:
            row_data[name_col] = wp.get('name', '')  # 第N職場名
        row_data[12] = ''                       # M: 初回パスワード（空）

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = data_font
            cell.alignment = data_align
            cell.border    = thin_border

    # 列幅の調整
    col_widths = [14, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16, 16]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 30  # ヘッダー行の高さ

    # バイト列として出力
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    output = make_response(buf.read())
    output.headers["Content-Disposition"] = "attachment; filename=workplaces.xlsx"
    output.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return output

# ----------------- CSV出力API（回答データ・指定フォーマット版） -----------------
@app.route('/api/csv')
def export_results_csv():
    responses = _db_list_responses()

    si = io.StringIO()
    cw = csv.writer(si)
    
    # 1. いただいたCSVフォーマットに合わせたヘッダーを作成
    headers = [
        '氏名', 'フリガナ', '生年月日(西暦)', '性別', '社員ID', '職場コード', '職場名',
        '前回社員ID', 'メールアドレス', '電話番号', '内線番号', '備考', '管理者コメント', '変数値'
    ]
    headers += [f'A-{i}' for i in range(1, 18)]
    headers += [f'B-{i}' for i in range(1, 30)]
    headers += [f'C-{i}' for i in range(1, 10)]
    headers += [f'D-{i}' for i in range(1, 3)]
    
    cw.writerow(headers)

    # 2. データをフォーマットに当てはめる
    for r in responses:
        try:
            ans = json.loads(r.get('answers_json') or '{}')
        except Exception:
            ans = {}

        row = [
            r.get('name', ''),              # 氏名
            ans.get('furigana', ''),        # フリガナ
            ans.get('birth_date', ''),      # 生年月日(西暦)
            ans.get('gender', ''),          # 性別
            r.get('employee_id', ''),       # 社員ID
            ans.get('workplace_code', ''),  # 職場コード
            r.get('workplace_name', ''),    # 職場名
            '',                             # 前回社員ID
            r.get('email') or '',           # メールアドレス
            '', '', '', '', ''              # 電話番号 〜 変数値までは空欄
        ]
        
        # A-1 〜 A-17 (q1 〜 q17)
        for i in range(1, 18):
            row.append(ans.get(f'q{i}', ''))
            
        # B-1 〜 B-29 (q18 〜 q46)
        for i in range(18, 47):
            row.append(ans.get(f'q{i}', ''))
            
        # C-1 〜 C-9 (q47 〜 q55)
        for i in range(47, 56):
            row.append(ans.get(f'q{i}', ''))
            
        # D-1 〜 D-2 (q56 〜 q57)
        for i in range(56, 58):
            row.append(ans.get(f'q{i}', ''))
            
        cw.writerow(row)

    # 取込システムで文字化けしないように「cp932 (Shift_JIS)」で出力します
    output = make_response(si.getvalue().encode('cp932', errors='replace'))
    output.headers["Content-Disposition"] = "attachment; filename=stress_results.csv"
    output.headers["Content-type"] = "text/csv"
    return output

# ==========================================
# サーバー起動設定（Gcloud Run対応）
# ==========================================
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=True)